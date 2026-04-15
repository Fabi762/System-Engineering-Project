import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import os, time
file_path = r"C:\Users\a932012\OneDrive - ATOS\Desktop\workspace_Uni\Requirements_Planung_Schaetzung_AKTUELL.xlsx"

# ============================================================================
# STYLES
# ============================================================================
thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
subheader_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
subheader_font = Font(bold=True, color="FFFFFF", size=10)

fertig_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
fertig_status_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
fertig_font = Font(color="065F46", bold=True)
fertig_status_font = Font(color="FFFFFF", bold=True)

offen_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
offen_status_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
offen_font = Font(color="92400E", bold=True)
offen_status_font = Font(color="FFFFFF", bold=True)

verschoben_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
verschoben_status_fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
verschoben_font = Font(color="991B1B", bold=True)

section_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
section_font = Font(bold=True, color="3730A3", size=11)

bold_font = Font(bold=True)
wrap_align = Alignment(wrap_text=True, vertical="top")
center_align = Alignment(horizontal="center", vertical="center")
center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ============================================================================
# SHEET 1: Milestone-Übersicht
# ============================================================================
ms_data = {
    "Milestone": [
        "M1: Projektsetup & Grundarchitektur",
        "M2: Dokumenten-Upload & Parsing",
        "M3: KI-Integration (Lernzettel)",
        "M4: Karteikarten & UI-Polish (geplant)"
    ],
    "Status": ["✅ Fertig", "✅ Fertig", "✅ Fertig", "⏳ Offen"],
    "Sprints": ["Sprint 1–2\n(Woche 1–2)", "Sprint 3–4\n(Woche 3–4)", "Sprint 5–6\n(Woche 5–6)", "Sprint 7–8\n(Woche 7–8)"],
    "Verantwortlich": ["Leon, Mario, Fabian", "Leon, Mario, Fabian", "Mario, Fabian, Leon", "Leon, Mario, Fabian"],
    "SP Geplant\n(Alte Planung)": [13, 21, 21, 10],
    "SP Tatsächlich /\nNeue Planung": [13, 19, 13, 39],
    "Abweichung\n(SP)": [0, -2, -8, +29],
    "Abweichung\n(%)": ["0%", "-10%", "-38%", "+290%"],
    "Begründung / Kommentar": [
        "Wie geplant abgeschlossen. Grundgerüst steht stabil.",
        "Docling-Integration lief schneller als erwartet. 2 SP eingespart.",
        "Karteikarten-Feature auf M4 verschoben → 8 SP weniger.\nPrompt Engineering war aufwändiger als gedacht.",
        "Enthält jetzt verschobene Features aus M3 (+21 SP),\nplus neue Anforderungen: Docker, Tests, Rate Limiting."
    ]
}
df_milestones = pd.DataFrame(ms_data)

# ============================================================================
# SHEET 2: User Stories & Story Points Vergleich
# ============================================================================
stories = {
    "ID": [
        "", "US-01", "US-08", "",
        "", "US-02", "US-06", "US-07", "US-REST-01", "",
        "", "US-04", "US-KI-01", "",
        "", "US-03", "US-05", "US-UI-01", "US-UI-02", "US-DEV-01", "US-DEV-02", ""
    ],
    "User Story / Feature": [
        "── MILESTONE 1: Projektsetup & Grundarchitektur ──",
        "Git-Repo, React+Vite, FastAPI aufsetzen, Projektstruktur",
        "Vite Dev Proxy konfigurieren (/api/* → :8000)",
        "",
        "── MILESTONE 2: Dokumenten-Upload & Parsing ──",
        "Drag & Drop Upload + Dateiauswahl mit Fortschrittsanzeige",
        "Seitenleiste: Dokumentenverwaltung (hinzufügen, wechseln, löschen)",
        "Markdown-Text kopieren (Copy-Button mit Feedback)",
        "Docling-Integration: PDF, PPTX, DOCX, XLSX, HTML, MD, Bilder",
        "",
        "── MILESTONE 3: KI-Integration (Lernzettel) ──",
        "KI-Lernzettel generieren + PDF-Export (Azure OpenAI GPT-4.1)",
        "Azure OpenAI Anbindung (.env, Endpoint, API-Version)",
        "",
        "── MILESTONE 4: Karteikarten & UI-Polish (GEPLANT) ──",
        "KI-Karteikarten generieren (Backend, JSON-Format, Prompt)",
        "Interaktive Karteikarten-UI (3D-Flip, Navigation, Fortschritt)",
        "Tab-Navigation: Wechsel zwischen Lernzettel ↔ Karteikarten",
        "Responsives Design (Mobile-Friendly, Design-System verfeinern)",
        "Docker-Compose + Unit Tests (pytest, Entwicklungsumgebung)",
        "Fehlerhandling & Rate Limiting (Azure OpenAI Stabilität)",
        ""
    ],
    "Status": [
        "", "✅ Fertig", "✅ Fertig", "",
        "", "✅ Fertig", "✅ Fertig", "✅ Fertig", "✅ Fertig", "",
        "", "✅ Fertig", "✅ Fertig", "",
        "", "⏳ Offen", "⏳ Offen", "⏳ Offen", "⏳ Offen", "⏳ Offen", "⏳ Offen", ""
    ],
    "Verantwortlich": [
        "", "Leon / Mario / Fabian", "Leon", "",
        "", "Leon", "Leon", "Fabian", "Mario", "",
        "", "Fabian", "Mario", "",
        "", "Mario", "Leon", "Leon", "Fabian", "Fabian", "Mario", ""
    ],
    "Priorität": [
        "", "Hoch", "Mittel", "",
        "", "Hoch", "Mittel", "Niedrig", "Mittel", "",
        "", "Hoch", "Hoch", "",
        "", "Hoch", "Mittel", "Mittel", "Mittel", "Niedrig", "Mittel", ""
    ],
    "SP\n(Alte Schätzung)": [
        "", 8, 5, "",
        "", 5, 3, 1, 5, "",
        "", 8, 5, "",
        "", 8, 5, 3, 5, 0, 0, ""
    ],
    "SP\n(Tatsächlich/Neu)": [
        "", 8, 5, "",
        "", 5, 3, 1, 5, "",
        "", 8, 5, "",
        "", 13, 8, 3, 5, 5, 5, ""
    ],
    "Δ SP": [
        "", 0, 0, "",
        "", 0, 0, 0, 0, "",
        "", 0, 0, "",
        "", +5, +3, 0, 0, "+5 (neu)", "+5 (neu)", ""
    ],
    "Akzeptanzkriterien / Details": [
        "",
        "Repo erstellt, React 19 + Vite läuft, FastAPI /api/health antwortet",
        "Frontend /api/* Requests werden an Backend :8000 weitergeleitet",
        "",
        "",
        "Upload per Drag & Drop UND Dateiauswahl, Ladeanimation bei Verarbeitung",
        "Dokumentenliste in Sidebar, aktives Dokument hervorgehoben, Löschen möglich",
        "Copy-Button kopiert Markdown in Zwischenablage, visuelles Feedback",
        "Alle Formate werden erkannt und in Markdown konvertiert (Docling)",
        "",
        "",
        "KI generiert strukturierten Lernzettel aus Markdown, PDF-Download automatisch",
        "API-Key + Endpoint aus .env geladen, Fehler bei fehlender Config",
        "",
        "",
        "Mind. 10 Karteikarten pro Dokument, JSON {question, answer}",
        "3D-Flip-Animation, Vor/Zurück-Navigation, Fortschrittsbalken",
        "Tab-Buttons im Viewer-Header, Zustand bleibt beim Wechsel erhalten",
        "App funktioniert auf Tablet & Smartphone, keine horizontalen Scrollbars",
        "docker-compose.yml, pytest Backend-Tests, identische Dev-Umgebung",
        "Retry bei Rate Limit, User-Feedback bei API-Fehlern, Graceful Degradation",
        ""
    ],
    "Notizen & Learnings": [
        "",
        "Reibungslose Umsetzung, klare Aufgabenverteilung bewährt sich.",
        "Einfaches Setup, spart CORS-Konfiguration im Deployment.",
        "",
        "",
        "Drag-Counter-Pattern verhindert flackernde Drag-Zustände.",
        "Relative Zeitangaben ('Vor 5 Min.') verbessern UX erheblich.",
        "Triviales Feature, schnell umgesetzt.",
        "Docling (IBM) liefert exzellente Ergebnisse bei PDFs. PPTX manchmal ungenau.",
        "",
        "",
        "Prompt Engineering war der größte Aufwand.\nfpdf2 hat Limitierungen bei verschachteltem Markdown.\nMAX_CONTENT_LENGTH (15k Zeichen) schneidet lange Vorlesungen ab.",
        "Azure-Doku teilweise veraltet. Mehrere Versuche für richtige API-Version nötig.",
        "",
        "",
        "Verschoben aus M3. Prompt muss für Q&A optimiert werden (anders als Lernzettel).",
        "Verschoben aus M3. CSS preserve-3d + backface-visibility nötig.",
        "Neues Feature für M4. useState für aktiven Tab.",
        "Aktuell nur Desktop. Media-Queries + Sidebar-Toggle für Mobile nötig.",
        "Neues Feature (Team-Wunsch aus Retrospektive). Bisher kein Docker/Tests.",
        "Neues Feature. Azure hat strikte Rate Limits. Ohne Retry bricht Generation ab.",
        ""
    ]
}
df_stories = pd.DataFrame(stories)

# ============================================================================
# SHEET 3: Sprint-Planung Detail
# ============================================================================
sprints = {
    "Sprint": [
        "Sprint 1", "Sprint 1", "Sprint 1", "Sprint 1", "Sprint 1",
        "Sprint 2", "Sprint 2", "Sprint 2", "Sprint 2", "Sprint 2", "Sprint 2",
        "Sprint 3", "Sprint 3", "Sprint 3",
        "Sprint 4 (geplant)", "Sprint 4 (geplant)", "Sprint 4 (geplant)", "Sprint 4 (geplant)", "Sprint 4 (geplant)", "Sprint 4 (geplant)"
    ],
    "Milestone": [
        "M1", "M1", "M1", "M1", "M1",
        "M2", "M2", "M2", "M2", "M2", "M2",
        "M3", "M3", "M3",
        "M4", "M4", "M4", "M4", "M4", "M4"
    ],
    "Aufgabe": [
        "Git-Repository einrichten", "React-Projekt mit Vite erstellen", "FastAPI-Backend aufsetzen",
        "Projektstruktur festlegen", "CI/CD Pipeline (optional)",
        "Drag & Drop Upload", "Docling-Integration", "Dokument-Ansicht (Markdown)",
        "Seitenleiste Dokumentenliste", "API-Endpunkte (CRUD)", "Fehlerbehandlung",
        "OpenAI-API Anbindung", "Lernzettel-Generierung + PDF", "Lernzettel-UI + Download",
        "Karteikarten-Generierung (Backend)", "Karteikarten-UI (Frontend)", "Tab-Navigation",
        "Responsives Design", "Docker + Tests", "Fehlerhandling + Rate Limiting"
    ],
    "Verantwortlich": [
        "Fabian", "Leon", "Mario", "Alle", "Fabian",
        "Leon", "Mario", "Fabian", "Leon", "Mario", "Fabian",
        "Mario", "Fabian", "Leon",
        "Mario", "Leon", "Leon", "Fabian", "Fabian", "Mario"
    ],
    "SP": [3, 3, 3, 2, 2, 5, 5, 3, 3, 5, 3, 5, 8, 5, 13, 8, 3, 5, 5, 5],
    "Status": [
        "Erledigt", "Erledigt", "Erledigt", "Erledigt", "Verschoben",
        "Erledigt", "Erledigt", "Erledigt", "Erledigt", "Erledigt", "Erledigt",
        "Erledigt", "Erledigt", "Erledigt",
        "Geplant", "Geplant", "Geplant", "Geplant", "Geplant", "Geplant"
    ]
}
df_sprints = pd.DataFrame(sprints)

# ============================================================================
# SHEET 4: Zusammenfassung / KPIs
# ============================================================================
summary = {
    "Metrik": [
        "Anzahl Milestones gesamt",
        "Milestones abgeschlossen",
        "Milestones offen",
        "Fortschritt (Milestones)",
        "",
        "Story Points gesamt (Alte Planung)",
        "Story Points gesamt (Neue Planung)",
        "Story Points abgeschlossen",
        "Story Points offen",
        "Fortschritt (Story Points)",
        "",
        "User Stories gesamt",
        "User Stories abgeschlossen",
        "User Stories offen",
        "",
        "Team-Velocity (Durchschnitt SP/Sprint)",
        "Geplante Velocity Sprint 4",
        "",
        "Größte Risiken für M4",
        "",
        ""
    ],
    "Wert": [
        4,
        3,
        1,
        "75%",
        "",
        65,
        84,
        45,
        39,
        "54%",
        "",
        11,
        7,
        4,
        "",
        "15 SP",
        "39 SP (ambitioniert!)",
        "",
        "Karteikarten-Prompt unklar, Scope zu groß",
        "",
        ""
    ],
    "Kommentar": [
        "",
        "M1, M2, M3",
        "M4 (Sprint 7–8)",
        "Auf Milestone-Ebene gut im Zeitplan",
        "",
        "Ursprüngliche Gesamtplanung über alle Milestones",
        "+19 SP durch verschobene Features und neue Anforderungen",
        "Alle Features aus M1–M3 erfolgreich implementiert",
        "Karteikarten (21 SP) + DevOps (10 SP) + UI (8 SP)",
        "Bezogen auf neue Gesamtplanung",
        "",
        "Inkl. neu hinzugefügte Stories (Docker, Rate Limiting)",
        "US-01 bis US-08 + US-KI-01, US-REST-01",
        "US-03, US-05, US-DEV-01, US-DEV-02 + UI Stories",
        "",
        "Basierend auf M1–M3 (45 SP / 3 Sprints)",
        "⚠ Deutlich über bisheriger Velocity – Scope ggf. reduzieren!",
        "",
        "Empfehlung: Karteikarten priorisieren, DevOps nur wenn Zeit bleibt",
        "",
        ""
    ]
}
df_summary = pd.DataFrame(summary)

# ============================================================================
# EXCEL SCHREIBEN
# ============================================================================
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    df_milestones.to_excel(writer, sheet_name='1. Milestones Übersicht', index=False)
    df_stories.to_excel(writer, sheet_name='2. Story Points Vergleich', index=False)
    df_sprints.to_excel(writer, sheet_name='3. Sprint-Planung Detail', index=False)
    df_summary.to_excel(writer, sheet_name='4. Zusammenfassung & KPIs', index=False)
    
    workbook = writer.book
    
    # ---- Formatierung für ALLE Sheets ----
    for sheet_name in writer.sheets:
        ws = writer.sheets[sheet_name]
        
        # Header-Zeile stylen
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_wrap
            cell.border = thin_border
        
        # Alle Zellen: Border + Wrap
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Spaltenbreite auto-anpassen (mit Limits)
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        lines = str(cell.value).split('\n')
                        for line in lines:
                            max_len = max(max_len, len(line))
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 55)
    
    # ---- Sheet 1: Milestones Übersicht – Status-Farben ----
    ws1 = writer.sheets['1. Milestones Übersicht']
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
        status_cell = row[1]  # Spalte B = Status
        is_fertig = "Fertig" in str(status_cell.value)
        
        for cell in row:
            if is_fertig:
                cell.fill = fertig_fill
            else:
                cell.fill = offen_fill
        
        if is_fertig:
            status_cell.fill = fertig_status_fill
            status_cell.font = fertig_status_font
        else:
            status_cell.fill = offen_status_fill
            status_cell.font = offen_status_font
        status_cell.alignment = center_align
    
    # ---- Sheet 2: Story Points Vergleich – Section Headers + Status ----
    ws2 = writer.sheets['2. Story Points Vergleich']
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        story_cell = row[1]  # Spalte B = User Story
        status_cell = row[2] # Spalte C = Status
        id_cell = row[0]     # Spalte A = ID
        
        # Section Header Zeilen (Milestone-Trenner)
        if str(story_cell.value or "").startswith("──"):
            for cell in row:
                cell.fill = section_fill
                cell.font = section_font
                cell.alignment = Alignment(vertical="center")
            continue
        
        # Leere Trennzeilen
        if not str(id_cell.value or "").strip():
            continue
        
        # Status-Zellen einfärben
        if "Fertig" in str(status_cell.value or ""):
            status_cell.fill = fertig_status_fill
            status_cell.font = fertig_status_font
            status_cell.alignment = center_align
        elif "Offen" in str(status_cell.value or ""):
            status_cell.fill = offen_status_fill
            status_cell.font = offen_status_font
            status_cell.alignment = center_align
    
    # ---- Sheet 3: Sprint-Planung – Status-Farben ----
    ws3 = writer.sheets['3. Sprint-Planung Detail']
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
        status_cell = row[5]  # Spalte F = Status
        status_val = str(status_cell.value or "")
        
        if status_val == "Erledigt":
            status_cell.fill = fertig_status_fill
            status_cell.font = fertig_status_font
        elif status_val == "Geplant":
            status_cell.fill = offen_status_fill
            status_cell.font = offen_status_font
        elif status_val == "Verschoben":
            status_cell.fill = verschoben_status_fill
            status_cell.font = Font(color="FFFFFF", bold=True)
        status_cell.alignment = center_align
    
    # ---- Sheet 4: Zusammenfassung – Hervorhebungen ----
    ws4 = writer.sheets['4. Zusammenfassung & KPIs']
    for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row):
        metrik = str(row[0].value or "")
        if "Fortschritt" in metrik or "Velocity" in metrik or "Risiken" in metrik:
            for cell in row:
                cell.font = bold_font
        if "75%" in str(row[1].value or ""):
            row[1].fill = fertig_status_fill
            row[1].font = fertig_status_font
            row[1].alignment = center_align
        if "ambitioniert" in str(row[1].value or ""):
            row[1].fill = offen_status_fill
            row[1].font = offen_status_font
    
    # Zeilenhöhe für bessere Lesbarkeit
    for sheet_name in writer.sheets:
        ws = writer.sheets[sheet_name]
        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 30

print(f"✅ Excel-Datei '{file_path}' wurde erfolgreich aktualisiert.")
print(f"   → Sheet 1: Milestones Übersicht (4 Milestones, Status-Ampel)")
print(f"   → Sheet 2: Story Points Vergleich (11 Stories, Alt vs. Neu)")
print(f"   → Sheet 3: Sprint-Planung Detail (20 Aufgaben)")
print(f"   → Sheet 4: Zusammenfassung & KPIs (Velocity, Fortschritt, Risiken)")
